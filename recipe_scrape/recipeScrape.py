from bs4 import BeautifulSoup
import urllib2
import re
from openpyxl import Workbook
from openpyxl import load_workbook

# specify url of main page to scrape

xl_book = 'drinks.xlsx'
drink_url = 'https://www.thespruceeats.com/a-to-z-cocktail-recipes-3962886'

def get_urls():
    urls = []
    # get html into page
    page = urllib2.urlopen(drink_url)

    # parse html and store in soup
    soup = BeautifulSoup(page, 'html.parser')

    # create excel workbook and worksheet
    wb = Workbook()
    ws = wb.active

    # each div contains a few recipes plus some extra data we don't need.
    divs = soup.findAll("div", id=re.compile("mntl-sc-block_.*"))

    # in each div, the unordered list contains all the drink links
    row = 1
    for div in divs:
        ul = div.find('ul') # each ul contains a set of drinks
        if ul: # ul error check
            for li in ul.findAll('li'): # each li is a drink
                a = li.find('a') # get link to drink recipe
                if a: # a error check
                    ws.cell(row = row, column = 1).value = a.text
                    this_url = a.get('href')
                    ws.cell(row = row, column = 2).value = this_url
                    urls.append(this_url)
                    row = row + 1
    wb.save(xl_book)
    print("Finished gathering urls")
    return urls

def get_details(urls):
    wb = load_workbook(filename = xl_book)
    ws = wb.active
    row = 1
    for url in urls:
        column = 3
        page = urllib2.urlopen(url)
        soup = BeautifulSoup(page, 'html.parser')
        # get all ingredient list items
        ingredients_li = soup.findAll("li", "simple-list__item js-checkbox-trigger ingredient")
        # add each item into worksheet
        for item in ingredients_li:
            if item is not None:
                #TODO: Deal with unusual amounts (fractions, stupid units, etc.)
                #TODO: Deal with Garnish type
                #TODO: Strip weird symbol that appears before some ingredient names
                full_string = item.text.strip('\n')
                if "Optional:" in full_string:
                    amount = "Optional"
                    item_name = full_string[10:]
                else: # not optional ingredient
                    split_string = full_string.split()
                    amount = ' '.join(split_string[:2])
                    item_name = ' '.join(split_string[2:])
                ws.cell(row = row, column = column).value = amount
                column = column + 1
                ws.cell(row = row, column = column).value = item_name
                column = column + 1
        # next url, increment sheet row
        row = row + 1
        print(url)
    wb.save(xl_book)
    return

def get_url_detail(url):
    wb = load_workbook(filename = xl_book)
    ws = wb.active
    row = 1
    column = 3
    ws.cell(3,3).value = 'test'
    page = urllib2.urlopen(url)
    soup = BeautifulSoup(page, 'html.parser')
    ingredients_li = soup.findAll("li", "simple-list__item js-checkbox-trigger ingredient")
    for item in ingredients_li:
        if item is not None:
            full_string = item.text.strip('\n')
            if "Optional:" in full_string:
                amount = "Optional"
                item_name = full_string[10:]
            else: # not optional ingredient
                split_string = full_string.split()
                amount = ' '.join(split_string[:2])
                item_name = ' '.join(split_string[2:])
            ws.cell(row = row, column = column).value = amount
            column = column + 1
            ws.cell(row = row, column = column).value = item_name
            column = column + 1
    wb.save(xl_book)
    return

def main():
    urls = get_urls()
    get_details(urls)
    # get_url_detail('https://www.thespruceeats.com/adult-hot-chocolate-recipe-759570')
    return

if __name__== "__main__":
    main()
