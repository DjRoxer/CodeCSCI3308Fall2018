# INSERT INTO drinks (name, ingredients, instructions)
# VALUES ('test drink',
#        ARRAY[['1 Ounce','Whiskey'],['2 Ounce','Rum']],
#        ARRAY['First instruction','Second Instruction','3rd instruction']
# );

# fill heroku db table drinks with those scraped into drinks.xlsx

from openpyxl import Workbook
from openpyxl import load_workbook
import psycopg2

xl_book = 'drinks.xlsx'
DATABASE_URL = 'postgres://fftbsmtlkjrdda:1f1ea62929ba0acbcbc74d4cd2267c04b70b1a2dfc1106c87aa440ec2ba70ee8@ec2-54-243-46-32.compute-1.amazonaws.com:5432/d7ncm4tt5a8edo'

# connect to postgers db
conn = psycopg2.connect(DATABASE_URL)
cur = conn.cursor()

# open spreadsheet
wb = load_workbook(filename = xl_book)
ws = wb.active
max_row = ws.max_row

# could use copy_from() to insert all the data at once, but i'm scared.

for i in range(1,max_row+1): # for each row of spreadsheet
    # get details of drink
    name = ws.cell(row=i,column=1).value
    link = ws.cell(row=i,column=2).value
    ingredients = []

    # fill ingredients array
    j = 3
    cell = ws.cell(row=i,column=j)
    while cell.value is not None:
        amount = cell.value
        ing = ws.cell(row=i,column=j+1).value
        ingredients.append([amount,ing])
        j = j + 2       # next ingredient amount
        cell = ws.cell(row=i,column=j)

    # add row to table
    cur.execute(
        "INSERT INTO drinks (name, ingredients, link) VALUES (%s, %s, %s)",
        (name, ingredients, link)
    )

conn.commit()
