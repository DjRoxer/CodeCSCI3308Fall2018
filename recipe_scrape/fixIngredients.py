#!/usr/bin/env python
# -*- coding: utf-8 -*-

import re
from openpyxl import Workbook
from openpyxl import load_workbook

xl_book = 'drinks.xlsx'

wb = load_workbook(filename = xl_book)
ws = wb.active

max_row = ws.max_row

# fix "ounces" in ingredient name

for i in range(1,max_row+1):  # for each row
        j = 4               # start at 4th column (first ingredient)
        cell = ws.cell(row=i,column=j)
        while cell.value is not None:
            if 'ounces' in cell.value:
                cell.value = cell.value[7:] # slice off ounces
                new_val = ws.cell(row=i,column=j-1).value + " ounces" # add ounces to amount
                ws.cell(row=i,column=j-1).value = new_val
            j = j + 2   # go to next column
            cell = ws.cell(row=i,column=j) # update cell to check

# remove zero-width space from offending cells

for i in range(1,max_row+1):
    j = 3
    cell = ws.cell(row=i,column=j)
    while cell.value is not None:
        str = cell.value
        str = str.replace(u'​','') # that first blank space is actually a zero-width space
        cell.value = str
        j = j + 1
        cell = ws.cell(row=i,column=j)

# fix Garnish

for i in range(1,max_row+1):
    j = 3
    cell = ws.cell(row=i,column=j)
    while cell.value is not None:
        if "Garnish:" in cell.value:
            stuff = cell.value[9:] # get stuff after "Garnish: "
            if stuff is not None:
                ing = ws.cell(row=i,column=j+1).value # current ingredient name
                if ing is not None:
                    ws.cell(row=i,column=j+1).value = stuff + ' ' + ing # prepend stuff after "Garnish: " to ing name
            cell.value = "Garnish" # fix amount to "Garnish"
        j = j + 2 # go to next ingredient amount
        cell = ws.cell(row=i,column=j)

wb.save(xl_book)
